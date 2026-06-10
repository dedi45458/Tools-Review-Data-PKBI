import streamlit as st

# TAMBAHKAN KODE INI DI BARIS ATAS
if 'total_entri' not in st.session_state:
    st.session_state['total_entri'] = 0

import streamlit as st
import pandas as pd
import io
import re
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ==========================================================
# 1. KONFIGURASI UTAMA & THEME MODERN
# ==========================================================
st.set_page_config(
    page_title="Data Quality Review - PKBI Jabar", 
    page_icon="📊", 
    layout="wide"
)

# Custom CSS untuk tampilan UI yang lebih bersih dan profesional
st.markdown("""
    <style>
    .main-title {
        font-size: 2.2rem;
        font-weight: 700;
        color: #1E3A8A;
        margin-bottom: 0.2rem;
    }
    .sub-title {
        font-size: 1.1rem;
        color: #4B5563;
        margin-bottom: 1.5rem;
    }
    .stButton>button {
        border-radius: 8px;
        font-weight: 600;
    }
    div[data-testid="stExpander"] {
        border-radius: 8px;
        border: 1px solid #E5E7EB;
    }
    </style>
    """, unsafe_allow_html=True)

st.markdown('<div class="main-title">📊 Tools Review Data Massal — PKBI Jabar</div>', unsafe_allow_html=True)
st.markdown('<div class="sub-title">Sistem otomatisasi penelaahan kualitas data Penjangkauan dan Rujukan PKBI Jawa Barat berbasis matriks validasi terbaru.</div>', unsafe_allow_html=True)

# Inisialisasi State di bagian atas agar tidak memicu KeyError saat rerun otomatis
if 'proses_selesai' not in st.session_state:
    st.session_state['proses_selesai'] = False
if 'data_unduhan' not in st.session_state:
    st.session_state['data_unduhan'] = None
if 'tabel_1_detail' not in st.session_state:
    st.session_state['tabel_1_detail'] = None
if 'tabel_2_matrik' not in st.session_state:
    st.session_state['tabel_2_matrik'] = None
if 'total_entri' not in st.session_state:
    st.session_state['total_entri'] = 0

# MASTER LIST 50 INDIKATOR KESALAHAN
DAFTAR_INDIKATOR = [
    "Tahun dalam tanggal penjangkauan lebih besar/kecil dari tahun sekarang",
    "Kode Petugas Kosong",
    "Tanggal lebih besar dari tanggal hari ini",
    "IDKD kurang/lebih dari 10 digit karakter",
    "Digit nama kurang/lebih dari 4 digit karakter",
    "Digit tanggal lahir lebih/kurang dari 6 digit angka",
    "Ada tanda titik (.) pada penulisan IDKD",
    "Ada spasi pada penulisan IDKD",
    "ID sama tapi NIK berbeda dengan data Semester/Tahun lalu (Konfirmasi)",
    "NIK sama tapi ID berbeda dengan data Semester/Tahun lalu (Konfirmasi)",
    "Tahun Lahir KD terlalu muda (2014 -sekarang)",
    "Usia KD dibawah 16 tahun (konfirmasi)",
    "Usia KD diatas 70 tahun (konfirmasi)",
    "Tahun lahir pada IDKD berbeda dengan Tahun lahir pada NIK (konfirmasi)",
    "NIK kurang/lebih dari 16 digit (konfirmasi)",
    "Kesalahan dalam penulisan NIK (00) (konfirmasi)",
    "Secara NIK harusnya perempuan bukan laki-laki (konfirmasi)",
    "LSL/Waria tapi jenis kelamin perempuan",
    "Jenis kontak dengan Jenis Kegiatan tidak sesuai",
    "Jenis kontak Individual/kelompok tapi kolom Virtual dan Tatap Muka (VC1) tidak diisi",
    "Penjangkauan tatap muka tapi lokasi outreach diindikasi ada nama medsos",
    "Lokasi outreach diisi IDKD",
    "Lokasi outreach diindikasi kurang spesifik atau kurang detil (digit huruf <17 digit) (konfirmasi)",
    "Lokasi outreach indikasi diisi nomer HP",
    "Bukan PWID mendapatkan info 8 atau 9 (LASS, PTRM)",
    "LSL/TG/PWID menerima informasi PMTC (konfirmasi)",
    "Konfirmasi jumlah KIE yang diberikan adalah wajar",
    "Konfirmasi jumlah kondom yang diberikan adalah wajar",
    "Konfirmasi jumlah pelicin yang diberikan adalah wajar",
    "Konfirmasi jumlah jarum yang diberikan adalah wajar",
    "Konfirmasi jumlah alkohol SWAB yang diberikan adalah wajar",
    "VO tapi kolom Virtual dan Tatap Muka (VC1) diisi angka 1",
    "VO tapi lokasi outreach bukan nama medsos/kurang tepat mencatat nama aplikasi medsos",
    "VO tapi menyerahkan jarum",
    "VO menerima logistik selain KIE",
    "VO tapi nama akun /No. Hp tidak diisi",
    "Tidak ada informasi satupun yang diberikan / tidak diisi",
    "KD dikontak lebih dari 1x tapi tidak mendapat informasi HIV",
    "KD telah menerima layanan CBS tapi tidak ada informasi CBS",
    "KD ada rujukan PrEp di penjangkauan tapi tidak ada informasi PrEp",
    "KD telah menerima layanan PrEp tapi tidak ada rujukan PrEp di penjangkauan",
    "Logistik kosong (Konfirmasi)",
    "Tipe klien PWID tapi tidak menerima jarum (konfirmasi)",
    "Tipe klien PWID tapi tidak menerima alkohol SWAB (konfirmasi)",
    "Popkun selain PWID menerima jarum suntik",
    "Popkun selain PWID menerima alkohol swab",
    "Popkun selain PWID menyerahkan jarum",
    "Tidak ada rujukan yang diberikan satupun / tidak diisi",
    "KD dikontak lebih dari 1x tetapi tidak ada Rujukan Tes HIV",
    "Bukan penasun rujukan 3,4"
]

# ==========================================================
# 2. PANEL SIDEBAR UNTUK UNGGAH BERKAS (UI ENHANCED)
# ==========================================================
with st.sidebar:
    st.markdown("### 📁 Menu Unggah Berkas")
    st.markdown("Pastikan format berkas sesuai standar laporan SSR.")
    
    file_referensi = st.file_uploader(
        "1️⃣ Data Semester / Tahun Lalu (.xlsx)", 
        type=["xlsx"],
        help="Digunakan untuk validasi silang kecocokan IDK Klien & NIK"
    )
    
    st.markdown("---")
    
    files_review = st.file_uploader(
        "2️⃣ Raw Data Penjangkauan (.xlsx)", 
        type=["xlsx"], 
        accept_multiple_files=True,
        help="Bisa memilih lebih dari 1 file sekaligus"
    )
    
    if files_review:
        st.success(f"📂 Terbaca {len(files_review)} file Penjangkauan.")
    else:
        st.info("💡 Silakan unggah file Penjangkauan untuk memulai analisis.")

def cek_kode(teks_kolom, kode_target):
    if pd.isna(teks_kolom): return False
    clean_str = str(teks_kolom).replace("'", "").replace(" ", "")
    list_kode = clean_str.split(",")
    return str(kode_target) in list_kode

# ==========================================================
# 3. ENGINE VALIDASI PENJANGKAUAN
# ==========================================================
def jalankan_review_penjangkauan(df_asli, df_ref=None, nama_file=""):
    list_kesalahan = []
    if df_asli.empty: return pd.DataFrame(list_kesalahan)
    
    df = df_asli.copy()
    df.columns = [str(c).strip() for c in df.columns]
    
    start_row_idx = 0
    if len(df) > 0 and ('dd/mm/yyyy' in str(df.iloc[0].values) or 'Laki-laki' in str(df.iloc[0].values)):
        start_row_idx = 1

    tahun_sekarang = datetime.now().year
    hari_ini = pd.Timestamp(datetime.now().date())
    
    medsoc_keywords = [
        'whatsapp', 'wa', 'badoo', 'hornet', 'michat', 'blued', 'bumble', 
        'walla', 'sms', 'grindr', 'growlr', 'instagram', 'ig', 'tantan', 
        'telegram', 'telepon', 'tinder', 'twitter', 'line', 'facebook', 'fb', 
        'messenger', 'romeo', 'tiktok', 'tagged', 'litmatch', 'scruff', 
        'wechat', 'threads'
    ]

    ref_ssr_id_to_nik = {}
    ref_nik_ssr_to_id = {}
    
    if df_ref is not None and not df_ref.empty:
        df_ref_cp = df_ref.copy()
        df_ref_cp.columns = [str(c).strip() for c in df_ref_cp.columns]
        col_id_ref = [c for c in df_ref_cp.columns if 'ID' in c or 'Klien' in c]
        col_nik_ref = [c for c in df_ref_cp.columns if 'NIK' in c]
        col_ssr_ref = [c for c in df_ref_cp.columns if 'SSR' in c or 'Lembaga' in c]
        
        if col_id_ref and col_nik_ref and col_ssr_ref:
            for _, r in df_ref_cp.iterrows():
                ssr_r = str(r[col_ssr_ref[0]]).strip().upper()
                id_r = str(r[col_id_ref[0]]).replace("'", "").strip()
                nik_r = str(r[col_nik_ref[0]]).replace("'", "").replace('.0', '').strip()
                
                if id_r and id_r != 'nan' and ssr_r and ssr_r != 'nan':
                    ref_ssr_id_to_nik[f"{ssr_r}_{id_r}"] = nik_r
                
                if nik_r and nik_r != 'nan' and nik_r != '' and ssr_r and ssr_r != 'nan':
                    ref_nik_ssr_to_id[f"{nik_r}_{ssr_r}"] = id_r

    id_counts = df.iloc[start_row_idx:]['ID Klien'].astype(str).str.strip().value_counts().to_dict()

    for idx in range(start_row_idx, len(df)):
        row = df.iloc[idx]
        no_excel_row = idx + 2
        
        v_ssr = str(row.get('Lembaga SSR', '')).strip().upper() if pd.notna(row.get('Lembaga SSR')) else ''
        v_petugas = str(row.get('Kode Petugas', '')).replace("'", "").strip() if pd.notna(row.get('Kode Petugas')) else ''
        v_kota = str(row.get('Nama Kota', '')).strip() if pd.notna(row.get('Nama Kota')) else ''
        v_tanggal = str(row.get('Tanggal', '')).split(' ')[0] if pd.notna(row.get('Tanggal')) else ''
        
        id_raw = str(row.get('ID Klien', '')).strip()
        id_clean = id_raw.replace("'", "").strip()
        
        nik_raw = str(row.get('NIK', '')).strip()
        nik_clean = nik_raw.replace("'", "").replace('.0', '').strip()

        v_tipe_sasaran = str(row.get('Tipe Sasaran', row.get('Tipe Klien', ''))).replace('.0', '').strip()
        umur = row.get('Umur', None)
        jk = str(row.get('Jenis Kelamin', '')).replace('.0', '').strip()
        jns_kontak = str(row.get('Jenis Kontak', '')).replace('.0', '').strip()
        jns_kegiatan = str(row.get('Jenis Kegiatan', '')).strip()
        lokasi = str(row.get('Lokasi Outreach / Jenis Sosial Media', '')).strip()
        info_diberikan = str(row.get('Informasi Yang diberikan', '')).strip()
        rujukan = str(row.get('Rujukan', '')).strip()
        no_hp = str(row.get('No. HP / Nama Akun', '')).strip()
        vc1 = str(row.get('Virtual & Tatap Muka', '')).replace('.0', '').strip()

        try:
            log_kie = float(row.iloc[17]) if pd.notna(row.iloc[17]) and str(row.iloc[17]).strip() not in ['', 'NaN'] else 0
            log_kon = float(row.iloc[18]) if pd.notna(row.iloc[18]) and str(row.iloc[18]).strip() not in ['', 'NaN'] else 0
            log_pel = float(row.iloc[19]) if pd.notna(row.iloc[19]) and str(row.iloc[19]).strip() not in ['', 'NaN'] else 0
            log_jar = float(row.iloc[20]) if pd.notna(row.iloc[20]) and str(row.iloc[20]).strip() not in ['', 'NaN'] else 0
            log_swab = float(row.iloc[21]) if pd.notna(row.iloc[21]) and str(row.iloc[21]).strip() not in ['', 'NaN'] else 0
            jarum_kembali = float(row.get('Jumlah Jarum Suntik Kembali', 0)) if pd.notna(row.get('Jumlah Jarum Suntik Kembali', 0)) else 0
        except:
            log_kie = log_kon = log_pel = log_jar = log_swab = jarum_kembali = 0

        tgl_raw = row.get('Tanggal', None)
        tgl_p = pd.to_datetime(tgl_raw, errors='coerce') if pd.notna(tgl_raw) else None

        def tambah_log(nama_indikator, deskripsi_detail):
            list_kesalahan.append({
                "Baris Excel": no_excel_row, 
                "Lembaga SSR": v_ssr,
                "Kode Petugas": v_petugas, 
                "Nama Kota": v_kota, 
                "Tanggal": v_tanggal, 
                "ID Klien": id_clean, 
                "NIK": nik_clean, 
                "Tipe Sasaran": v_tipe_sasaran,
                "INDIKATOR KESALAHAN DATA": nama_indikator
            })

        # --- LOGIKA VALIDASI INDIKATOR ---
        if pd.isna(row.get('Kode Petugas')) or str(row.get('Kode Petugas')).strip() == '':
            tambah_log(DAFTAR_INDIKATOR[1], "Kode petugas kosong")

        if pd.notna(tgl_p) and tgl_p > hari_ini:
            tambah_log(DAFTAR_INDIKATOR[2], "Tanggal pelaksanaan melebihi tanggal hari ini")

        if id_clean and id_clean != 'nan' and id_clean != '':
            if len(id_clean) != 10 or not id_clean.isalnum():
                tambah_log(DAFTAR_INDIKATOR[3], f"IDKD berjumlah {len(id_clean)} karakter")
            if len(id_clean) >= 4 and not id_clean[:4].isalpha():
                tambah_log(DAFTAR_INDIKATOR[4], f"4 digit awal IDKD wajib huruf")
            if len(id_clean) == 10 and not id_clean[4:].isdigit():
                tambah_log(DAFTAR_INDIKATOR[5], f"6 digit akhir IDKD wajib angka")
            
            # Pola Validasi No 9 Hanya jika Data Referensi diunggah
            if df_ref is not None and v_ssr:
                key_ssr_id = f"{v_ssr}_{id_clean}"
                if key_ssr_id in ref_ssr_id_to_nik and ref_ssr_id_to_nik[key_ssr_id] != nik_clean:
                    tambah_log(DAFTAR_INDIKATOR[8], f"ID terikat NIK berbeda dengan semester lalu")

        # Pola Validasi No 10 Hanya jika Data Referensi diunggah
        if df_ref is not None and v_ssr and nik_clean and nik_clean != 'nan' and nik_clean != '':
            key_nik_ssr = f"{nik_clean}_{v_ssr}"
            if key_nik_ssr in ref_nik_ssr_to_id and ref_nik_ssr_to_id[key_nik_ssr] != id_clean:
                tambah_log(DAFTAR_INDIKATOR[9], f"NIK terikat ID berbeda dengan semester lalu")

        if pd.notna(umur) and str(umur).strip() != '':
            try:
                val_umur = float(umur)
                if val_umur < 17: 
                    tambah_log(DAFTAR_INDIKATOR[11], f"Usia di bawah 17 tahun")
                if val_umur > 70: 
                    tambah_log(DAFTAR_INDIKATOR[12], f"Usia di atas 70 tahun")
            except: pass

        if id_clean and len(id_clean) == 10 and nik_clean and len(nik_clean) == 16:
            thn_id = id_clean[4:6]
            nik_with_quote = nik_raw if nik_raw.startswith("'") else "'" + nik_clean
            if len(nik_with_quote) >= 13:
                thn_nik = nik_with_quote[11:13]
                if thn_id != thn_nik:
                    tambah_log(DAFTAR_INDIKATOR[13], f"Tahun lahir IDKD tidak sama dengan Tahun lahir NIK")

        if nik_clean and nik_clean != 'nan' and nik_clean != '':
            if len(nik_clean) != 16:
                tambah_log(DAFTAR_INDIKATOR[14], f"NIK berjumlah {len(nik_clean)} digit")
            if nik_clean.endswith('00'):
                tambah_log(DAFTAR_INDIKATOR[15], "Penulisan NIK dummy")

        if len(nik_clean) == 16 and jk == '1':
            try:
                dd_nik = int(nik_clean[6:8])
                if dd_nik > 31:
                    tambah_log(DAFTAR_INDIKATOR[16], f"Klien Laki-laki (1) tapi NIK menunjukkan Perempuan")
            except: pass

        if (v_tipe_sasaran in ['1304', '1301']) and jk == '2': 
            tambah_log(DAFTAR_INDIKATOR[17], f"Populasi LSL/Waria tetapi Jenis Kelamin Perempuan")

        if jns_kontak == '1' and jns_kegiatan not in ['1', '5']:
            tambah_log(DAFTAR_INDIKATOR[18], f"Kontak Individual tidak sinkron dengan Kegiatan")
        elif jns_kontak == '2' and jns_kegiatan not in ['2', '3', '4', '6', '7']:
            tambah_log(DAFTAR_INDIKATOR[18], f"Kontak Kelompok tidak sinkron dengan Kegiatan")
        elif jns_kontak == '3' and jns_kegiatan != '8':
            tambah_log(DAFTAR_INDIKATOR[18], f"Kontak Virtual/VO tidak sinkron dengan Kegiatan")

        if '.' in id_raw: tambah_log(DAFTAR_INDIKATOR[6], "Ada tanda titik (.) pada penulisan IDKD")
        if ' ' in id_raw: tambah_log(DAFTAR_INDIKATOR[7], "Ada spasi pada penulisan IDKD")
        
        if pd.notna(umur) and str(umur).strip() != '':
            try:
                val_umur = float(umur)
                tahun_lahir = tahun_sekarang - val_umur
                if hasattr(tahun_lahir, 'year'):
                    tahun_lahir = tahun_lahir.year
                if 2014 <= tahun_lahir <= tahun_sekarang: 
                    tambah_log(DAFTAR_INDIKATOR[10], "Tahun lahir terlalu muda")
            except: pass

        is_vo = (jns_kontak == '3')
        any_medsoc_in_lokasi = any(kw in lokasi.lower() for kw in medsoc_keywords)

        if jns_kontak in ['1', '2']:
            if vc1 == '' or vc1 == 'nan': tambah_log(DAFTAR_INDIKATOR[19], "Tatap muka tapi VC1 kosong")
            if any_medsoc_in_lokasi: tambah_log(DAFTAR_INDIKATOR[20], "Tatap muka tapi lokasi ada nama medsos")

        if is_vo:
            if vc1 == '1': tambah_log(DAFTAR_INDIKATOR[31], "VO tapi VC1 diisi angka 1")
            if lokasi and not any_medsoc_in_lokasi: tambah_log(DAFTAR_INDIKATOR[32], "VO tapi lokasi outreach bukan nama medsos")
            if log_jar > 0: tambah_log(DAFTAR_INDIKATOR[33], "VO tapi menyerahkan jarum")
            if log_kon > 0 or log_pel > 0 or log_swab > 0: tambah_log(DAFTAR_INDIKATOR[34], "VO menerima logistik selain KIE")
            if no_hp == '' or no_hp == 'nan': tambah_log(DAFTAR_INDIKATOR[35], "VO tapi nama akun / No HP kosong")

        if lokasi and lokasi != 'nan':
            if len(lokasi) == 10 and lokasi[:4].isalpha() and lokasi[4:].isdigit(): tambah_log(DAFTAR_INDIKATOR[21], "Lokasi diisi IDKD")
            if len(lokasi) < 17 and not is_vo: tambah_log(DAFTAR_INDIKATOR[22], "Lokasi outreach kurang detil (<17 digit)")
            if re.search(r'(08\d{8,11})|(\+62\d{8,11})', lokasi.replace('-', '').replace(' ', '')): tambah_log(DAFTAR_INDIKATOR[23], "Lokasi diisi nomor HP")

        is_pwid = (v_tipe_sasaran == '1401')
        if not is_pwid:
            if cek_kode(info_diberikan, '8') or cek_kode(info_diberikan, '9') or cek_kode(jns_kegiatan, '8') or cek_kode(jns_kegiatan, '9'):
                tambah_log(DAFTAR_INDIKATOR[24], "Bukan PWID mendapatkan info LASS/PTRM")
            if log_jar > 0: tambah_log(DAFTAR_INDIKATOR[44], "Popkun selain PWID menerima jarum")
            if log_swab > 0: tambah_log(DAFTAR_INDIKATOR[45], "Popkun selain PWID menerima alkohol swab")
            if jarum_kembali > 0: tambah_log(DAFTAR_INDIKATOR[46], "Popkun selain PWID menyerahkan jarum")
            if cek_kode(rujukan, '3') or cek_kode(rujukan, '4'): tambah_log(DAFTAR_INDIKATOR[49], "Bukan penasun rujukan 3,4")
        else:
            if log_jar == 0 and not is_vo: tambah_log(DAFTAR_INDIKATOR[42], "Tipe klien PWID tapi tidak menerima jarum")
            if log_swab == 0 and not is_vo: tambah_log(DAFTAR_INDIKATOR[43], "Tipe klien PWID tapi tidak menerima alkohol swab")

        if (v_tipe_sasaran in ['1304', '1301', '1401']) and (cek_kode(info_diberikan, '6') or cek_kode(jns_kegiatan, '6')):
            tambah_log(DAFTAR_INDIKATOR[25], "LSL/TG/PWID menerima informasi PMTC")

        if log_kie > 10: tambah_log(DAFTAR_INDIKATOR[26], "KIE tidak wajar")
        if log_kon > 144: tambah_log(DAFTAR_INDIKATOR[27], "Kondom tidak wajar")
        if log_pel > 50: tambah_log(DAFTAR_INDIKATOR[28], "Pelicin tidak wajar")
        if log_jar > 100: tambah_log(DAFTAR_INDIKATOR[29], "Jarum tidak wajar")
        if log_swab > 100: tambah_log(DAFTAR_INDIKATOR[30], "Alkohol swab tidak wajar")

        if info_diberikan == '' or info_diberikan == 'nan': tambah_log(DAFTAR_INDIKATOR[36], "Tidak ada informasi satupun yang diberikan")
        if log_kie == 0 and log_kon == 0 and log_pel == 0 and log_jar == 0 and log_swab == 0: tambah_log(DAFTAR_INDIKATOR[41], "Logistik kosong")
        if rujukan == '' or rujukan == 'nan': tambah_log(DAFTAR_INDIKATOR[47], "Tidak ada rujukan yang diberikan satupun")

        if id_clean and id_counts.get(id_clean, 0) > 1:
            df_klien_ini = df[df['ID Klien'].astype(str).str.replace("'", "").str.strip() == id_clean]
            pernah_dapat_info_hiv = any(cek_kode(inf, '1') for inf in df_klien_ini['Informasi Yang diberikan'].values) or any(cek_kode(keg, '1') for keg in df_klien_ini['Jenis Kegiatan'].values)
            pernah_dapat_rujuk_tes = any(cek_kode(ruj, '2') for ruj in df_klien_ini['Rujukan'].values)
            if not pernah_dapat_info_hiv: tambah_log(DAFTAR_INDIKATOR[37], "KD dikontak lebih dari 1x tapi tidak mendapat info HIV")
            if not pernah_dapat_rujuk_tes: tambah_log(DAFTAR_INDIKATOR[48], "KD dikontak lebih dari 1x tetapi tidak ada Rujukan Tes HIV")

        if cek_kode(jns_kegiatan, '13') and not cek_kode(info_diberikan, '13'): tambah_log(DAFTAR_INDIKATOR[38], "KD telah menerima layanan CBS tapi tidak ada info CBS")
        if (cek_kode(rujukan, '5') or cek_kode(jns_kegiatan, '10')) and not (cek_kode(info_diberikan, '10') or cek_kode(jns_kegiatan, '10')):
            tambah_log(DAFTAR_INDIKATOR[39], "KD ada rujukan PrEp tapi tidak ada informasi PrEp")
        if cek_kode(jns_kegiatan, '10') and not cek_kode(rujukan, '5'): tambah_log(DAFTAR_INDIKATOR[40], "KD telah menerima layanan PrEp tapi tidak ada rujukan PrEp")

    return pd.DataFrame(list_kesalahan)

# ==========================================================
# 4. TOMBOL EKSEKUSI UTAMA
# ==========================================================
col_btn, _ = st.columns([1, 2])
with col_btn:
    tombol_proses = st.button("🚀 Jalankan Penelaahan Laporan", type="primary", use_container_width=True)

if tombol_proses:
    if not files_review:
        st.error("⚠️ **Gagal Mengeksekusi:** Silakan unggah berkas Raw Data terlebih dahulu di sidebar!")
    else:
        with st.spinner("Sedang memproses validasi data, mohon tunggu..."):
            
            MASTER_LIST_SSR = [
                "BINA MUDA GEMILANG", "YAYASAN PONTIANAK PLUS - OUTREACH", "YAYASAN PESONA BENGKULU", 
                "YAYASAN SRIKANDI PASUNDAN", "GRAPIKS", "LENSA SUKABUMI", "PETIK", 
                "PKBI CIREBON", "PKBI GARUT", "PKBI CABANG SUBANG", "YAYASAN SRIKANDI PERINTIS", 
                "PESONA BUMI PASUNDAN", "LEMBAGA KASIH INDONESIA KITA", "YAYASAN PELANGI MALUKU", 
                "YAYASAN VESTA INDONESIA", "WAHANA CITA INDONESIA"
            ]
            
            unique_ssrs = set(MASTER_LIST_SSR)
            total_records_processed = 0
            
            for file in files_review:
                try:
                    df_temp = pd.read_excel(file)
                    df_temp.columns = [str(c).strip() for c in df_temp.columns]
                    if 'Lembaga SSR' in df_temp.columns:
                        start_row = 1 if len(df_temp) > 0 and ('dd/mm/yyyy' in str(df_temp.iloc[0].values) or 'Laki-laki' in str(df_temp.iloc[0].values)) else 0
                        ssrs_in_file = df_temp.iloc[start_row:]['Lembaga SSR'].dropna().astype(str).str.strip().upper().unique()
                        for s in ssrs_in_file:
                            if s and s != 'nan' and s != '': 
                                unique_ssrs.add(s)
                        total_records_processed += (len(df_temp) - start_row)
                except Exception as e:
                    st.warning(f"Catatan: Masalah membaca file {file.name}: {e}")
                    
            list_ssr_unik = sorted(list(unique_ssrs))

            df_ref = None
            if file_referensi:
                try: 
                    df_ref = pd.read_excel(file_referensi)
                except Exception as e: 
                    st.warning(f"Gagal memuat Berkas Referensi Semester Lalu: {e}")
            
            semua_rekap_kesalahan = []
            for file in files_review:
                try:
                    df_target = pd.read_excel(file)
                    df_rekap_file = jalankan_review_penjangkauan(df_target, df_ref, nama_file=file.name)
                    if not df_rekap_file.empty:
                        semua_rekap_kesalahan.append(df_rekap_file)
                except Exception as e:
                    st.error(f"Gagal memproses file {file.name}: {e}")
            
            kolom_log = ['Baris Excel', 'Lembaga SSR', 'Kode Petugas', 'Nama Kota', 'Tanggal', 'ID Klien', 'NIK', 'Tipe Sasaran', 'INDIKATOR KESALAHAN DATA']

            if semua_rekap_kesalahan:
                df_tabel_1 = pd.concat(semua_rekap_kesalahan, ignore_index=True)
                df_tabel_1['Lembaga SSR'] = df_tabel_1['Lembaga SSR'].str.strip().str.upper()
                df_tabel_1 = df_tabel_1[kolom_log]
            else:
                df_tabel_1 = pd.DataFrame(columns=kolom_log)
            
            matrix_data = []
            for idx, ind in enumerate(DAFTAR_INDIKATOR, 1):
                row_dict = {"No.": idx, "INDIKATOR KESALAHAN DATA": ind}
                total_row_err = 0
                for ssr in list_ssr_unik:
                    if not df_tabel_1.empty:
                        count_err = len(df_tabel_1[(df_tabel_1['INDIKATOR KESALAHAN DATA'] == ind) & 
                                                  (df_tabel_1['Lembaga SSR'].str.upper() == ssr)])
                    else:
                        count_err = 0
                    row_dict[ssr] = count_err
                    total_row_err += count_err
                
                row_dict["Jumlah per indikator"] = total_row_err
                pct = (total_row_err / total_records_processed * 100) if total_records_processed > 0 else 0
                row_dict["% Kesalahan"] = f"{pct:.2f}%"
                matrix_data.append(row_dict)
                           
            df_tabel_2_matrik = pd.DataFrame(matrix_data)
            
            # --- OPENPYXL EXCEL GABUNGAN DESIGN ---
            output_stream = io.BytesIO()
            wb = openpyxl.Workbook()
            
            ws_dash = wb.active
            ws_dash.title = "Tabel 2 - Rekap Matrik SSR"
            ws_dash.views.sheetView[0].showGridLines = True
            
            font_header = Font(name="Calibri", size=9, bold=True)
            font_sec = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            font_data = Font(name="Calibri", size=10)
            
            fill_blue_header = PatternFill(start_color="BCD6EE", end_color="BCD6EE", fill_type="solid")
            fill_orange_summary = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
            fill_section_bar = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
            
            thin_border = Border(
                left=Side(style='thin', color='BFBFBF'), right=Side(style='thin', color='BFBFBF'),
                top=Side(style='thin', color='BFBFBF'), bottom=Side(style='thin', color='BFBFBF')
            )
            
            ws_dash['A1'] = "No."
            ws_dash['B1'] = "INDIKATOR KESALAHAN DATA"
            ws_dash.merge_cells('A1:A2')
            ws_dash.merge_cells('B1:B2')
            
            start_col_idx = 3
            end_col_idx = start_col_idx + len(list_ssr_unik) - 1
            ws_dash.cell(row=1, column=start_col_idx, value="Jml. Kesalahan data")
            if len(list_ssr_unik) > 1:
                ws_dash.merge_cells(start_row=1, start_column=start_col_idx, end_row=1, end_column=end_col_idx)
                
            for i, ssr in enumerate(list_ssr_unik):
                ws_dash.cell(row=2, column=start_col_idx + i, value=ssr)
                
            col_jml = end_col_idx + 1
            col_pct = end_col_idx + 2
            ws_dash.cell(row=1, column=col_jml, value="Jumlah per indikator")
            ws_dash.cell(row=1, column=col_pct, value="% Kesalahan")
            ws_dash.merge_cells(start_row=1, start_column=col_jml, end_row=2, end_column=col_jml)
            ws_dash.merge_cells(start_row=1, start_column=col_pct, end_row=2, end_column=col_pct)
            
            for r in [1, 2]:
                for c in range(1, col_pct + 1):
                    cell = ws_dash.cell(row=r, column=c)
                    cell.font = font_header
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.fill = fill_orange_summary if c in [col_jml, col_pct] else fill_blue_header
                    cell.border = thin_border

            ws_dash.cell(row=3, column=1, value="Penjangkauan")
            ws_dash.merge_cells(start_row=3, start_column=1, end_row=3, end_column=col_pct)
            ws_dash.cell(row=3, column=1).fill = fill_section_bar
            ws_dash.cell(row=3, column=1).font = font_sec
            ws_dash.cell(row=3, column=1).alignment = Alignment(vertical="center", indent=1)
            ws_dash.row_dimensions[3].height = 24
            
            for idx, ind in enumerate(DAFTAR_INDIKATOR):
                current_row = 4 + idx
                ws_dash.cell(row=current_row, column=1, value=idx + 1)
                ws_dash.cell(row=current_row, column=2, value=ind)
                
                total_row_err = 0
                for i, ssr in enumerate(list_ssr_unik):
                    if not df_tabel_1.empty:
                        count_err = len(df_tabel_1[(df_tabel_1['INDIKATOR KESALAHAN DATA'] == ind) & (df_tabel_1['Lembaga SSR'] == ssr)])
                    else:
                        count_err = 0
                    val_cell = count_err if count_err > 0 else "-"
                    c_cell = ws_dash.cell(row=current_row, column=start_col_idx + i, value=val_cell)
                    c_cell.alignment = Alignment(horizontal="center")
                    total_row_err += count_err
                
                j_cell = ws_dash.cell(row=current_row, column=col_jml, value=total_row_err if total_row_err > 0 else "-")
                j_cell.alignment = Alignment(horizontal="center")
                j_cell.fill = fill_orange_summary
                
                pct_val = (total_row_err / total_records_processed) if total_records_processed > 0 else 0
                p_cell = ws_dash.cell(row=current_row, column=col_pct, value=pct_val)
                p_cell.number_format = '0.00%'
                p_cell.alignment = Alignment(horizontal="center")
                p_cell.fill = fill_orange_summary
                
                for c in range(1, col_pct + 1):
                    cell = ws_dash.cell(row=current_row, column=c)
                    cell.font = font_data
                    cell.border = thin_border
            
            ws_dash.column_dimensions['B'].width = 65

            ws_detail = wb.create_sheet(title="Tabel 1 - Detail Per Baris")
            ws_detail.views.sheetView[0].showGridLines = True
            
            ws_detail.append(kolom_log)
            for c_idx, col_name in enumerate(kolom_log, 1):
                cell = ws_detail.cell(row=1, column=c_idx)
                cell.font = font_header
                cell.fill = fill_blue_header
                cell.border = thin_border
                
            if not df_tabel_1.empty:
                for r_data in df_tabel_1[kolom_log].values.tolist():
                    ws_detail.append(r_data)
                    for c_idx in range(1, len(kolom_log) + 1):
                        ws_detail.cell(row=ws_detail.max_row, column=c_idx).font = font_data
                        ws_detail.cell(row=ws_detail.max_row, column=c_idx).border = thin_border
            else:
                ws_detail.append(["-", "CLEAN", "-", "-", "-", "-", "-", "-", "Tidak ditemukan kesalahan data."])
            
            wb.save(output_stream)
            
            # Simpan hasil proses akhir ke session state
            st.session_state['data_unduhan'] = output_stream.getvalue()
            st.session_state['tabel_1_detail'] = df_tabel_1
            st.session_state['tabel_2_matrik'] = df_tabel_2_matrik
            st.session_state['total_entri'] = total_records_processed
            st.session_state['proses_selesai'] = True

# ==========================================================
# 5. BLOCK OUTPUT INTERFACE UTAMA STREAMLIT (UI MODERN)
# ==========================================================
if st.session_state['proses_selesai']:
    st.markdown("### 📊 Dashboard Hasil Review Analisis")
    
    # KARTU METRIK KINERJA (Modern Cards)
    m1, m2, m3 = st.columns(3)
    with m1:
        st.metric(label="Total Entri Data Diperiksa", value=f"{st.session_state['total_entri']} Baris")
    with m2:
        total_error = len(st.session_state['tabel_1_detail'])
        st.metric(label="Total Temuan Log Kesalahan", value=f"{total_error} Kasus", delta=f"{'⚠️ Perlu Tindakan' if total_error > 0 else '✅ Data Bersih'}")
    with m3:
        st.download_button(
            label="📥 Unduh File Excel Komplit (.xlsx)",
            data=st.session_state['data_unduhan'],
            file_name=f"Hasil_Review_Data_PKBI_Jabar_{datetime.now().strftime('%d%m%Y')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
        
    st.markdown("---")
    
    # STRUKTUR TAB PREVIEW INTERACTIVE
    tab1, tab2 = st.tabs(["📋 Tabel 2 - Matrik Ringkasan SSR", "🔍 Tabel 1 - Detail Log Kesalahan Per Baris"])
    
    with tab1:
        st.markdown("##### Ringkasan Distribusi Temuan Masalah per Lembaga SSR")
        if st.session_state['tabel_2_matrik'] is not None:
            # Hilangkan index bawaan pandas agar tampilan tabel bersih
            st.dataframe(st.session_state['tabel_2_matrik'].reset_index(drop=True), use_container_width=True)
            
    with tab2:
        st.markdown("##### Log Temuan Data Tidak Sinkron (Detail Per Baris)")
        if st.session_state['tabel_1_detail'] is not None and not st.session_state['tabel_1_detail'].empty:
            st.dataframe(st.session_state['tabel_1_detail'].reset_index(drop=True), use_container_width=True)
        else:
            st.success("✨ Luar biasa! Tidak ditemukan anomali atau kesalahan data dari berkas laporan yang diperiksa.")
